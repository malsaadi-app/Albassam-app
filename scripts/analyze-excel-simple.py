#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from collections import Counter
import json

file_path = '/data/.openclaw/workspace/albassam-tasks/data/employees_data.xlsx'

print("\n📊 تحليل ملف بيانات الموظفين:")
print("=" * 80)

try:
    # قراءة الملف
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    
    # قراءة العناوين
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    print(f"\n✅ تم قراءة الملف بنجاح!")
    print(f"📋 عدد الأعمدة: {len(headers)}")
    
    print("\n📑 أسماء الأعمدة:")
    for i, col in enumerate(headers, 1):
        print(f"   {i}. {col}")
    
    # قراءة كل الصفوف
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(row)
    
    print(f"\n📋 عدد الموظفين: {len(rows)}")
    
    # عرض عينة
    print("\n" + "=" * 80)
    print("\n🔍 عينة من البيانات (أول 3 صفوف):")
    for i, row in enumerate(rows[:3], 1):
        print(f"\n   صف {i}:")
        for header, value in zip(headers, row):
            if value:
                print(f"      {header}: {value}")
    
    # تحليل المسميات الوظيفية
    job_col_idx = None
    for i, header in enumerate(headers):
        if header and ('وظيف' in str(header) or 'مسمى' in str(header) or 'position' in str(header).lower()):
            job_col_idx = i
            break
    
    if job_col_idx is not None:
        print("\n" + "=" * 80)
        print(f"\n💼 المسميات الوظيفية (العمود: {headers[job_col_idx]}):")
        jobs = Counter()
        for row in rows:
            if row[job_col_idx]:
                jobs[str(row[job_col_idx]).strip()] += 1
        
        print(f"\n   إجمالي المسميات المختلفة: {len(jobs)}")
        print("\n   أكثر 30 مسمى تكراراً:")
        for job, count in jobs.most_common(30):
            print(f"      {count:3d}x - {job}")
        
        # حفظ كل المسميات
        with open('/data/.openclaw/workspace/albassam-tasks/data/job_titles.json', 'w', encoding='utf-8') as f:
            json.dump(dict(jobs), f, ensure_ascii=False, indent=2)
        print("\n   ✅ تم حفظ كل المسميات في: data/job_titles.json")
    
    # تحليل الأقسام
    dept_col_idx = None
    for i, header in enumerate(headers):
        if header and ('قسم' in str(header) or 'department' in str(header).lower()):
            dept_col_idx = i
            break
    
    if dept_col_idx is not None:
        print("\n" + "=" * 80)
        print(f"\n🏢 الأقسام (العمود: {headers[dept_col_idx]}):")
        depts = Counter()
        for row in rows:
            if row[dept_col_idx]:
                depts[str(row[dept_col_idx]).strip()] += 1
        
        print(f"\n   إجمالي الأقسام المختلفة: {len(depts)}")
        print("\n   القائمة الكاملة:")
        for dept, count in depts.most_common():
            print(f"      {count:3d}x - {dept}")
        
        # حفظ الأقسام
        with open('/data/.openclaw/workspace/albassam-tasks/data/departments.json', 'w', encoding='utf-8') as f:
            json.dump(dict(depts), f, ensure_ascii=False, indent=2)
        print("\n   ✅ تم حفظ الأقسام في: data/departments.json")
    
    # تحليل الفروع
    branch_col_idx = None
    for i, header in enumerate(headers):
        if header and ('فرع' in str(header) or 'branch' in str(header).lower() or 'مجمع' in str(header)):
            branch_col_idx = i
            break
    
    if branch_col_idx is not None:
        print("\n" + "=" * 80)
        print(f"\n🏭 الفروع (العمود: {headers[branch_col_idx]}):")
        branches = Counter()
        for row in rows:
            if row[branch_col_idx]:
                branches[str(row[branch_col_idx]).strip()] += 1
        
        print(f"\n   إجمالي الفروع المختلفة: {len(branches)}")
        print("\n   القائمة الكاملة:")
        for branch, count in branches.most_common():
            print(f"      {count:3d}x - {branch}")
        
        # حفظ الفروع
        with open('/data/.openclaw/workspace/albassam-tasks/data/branches.json', 'w', encoding='utf-8') as f:
            json.dump(dict(branches), f, ensure_ascii=False, indent=2)
        print("\n   ✅ تم حفظ الفروع في: data/branches.json")
    
    # حفظ ملخص عام
    summary = {
        "total_employees": len(rows),
        "total_columns": len(headers),
        "columns": headers,
        "job_titles_count": len(jobs) if job_col_idx else 0,
        "departments_count": len(depts) if dept_col_idx else 0,
        "branches_count": len(branches) if branch_col_idx else 0
    }
    
    with open('/data/.openclaw/workspace/albassam-tasks/data/summary.json', 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    
    print("\n" + "=" * 80)
    print("\n✅ تم الانتهاء من التحليل!")
    print("✅ الملفات المحفوظة:")
    print("   - data/summary.json (ملخص عام)")
    print("   - data/job_titles.json (المسميات الوظيفية)")
    print("   - data/departments.json (الأقسام)")
    print("   - data/branches.json (الفروع)")
    
    wb.close()
    
except Exception as e:
    print(f"\n❌ خطأ: {e}")
    import traceback
    traceback.print_exc()
